"""
Beer Label Analysis Pipeline
---------------------------
This script analyzes beer bottle/can images to extract:
- OCR text from the label
- Dominant colors of the label and the bottle/can
- The original beer URL (from beers.csv)

It outputs:
- A CSV file with the analysis results
- An Excel file with color-filled cells and clickable beer URLs

Assumptions:
- Images are in the 'beer_images' directory
- beers.csv contains the mapping from image file to beer URL
- Backgrounds are mostly plain/white
"""

import os
import cv2
import numpy as np
import pandas as pd
from sklearn.cluster import KMeans
import easyocr
import openpyxl
from openpyxl.styles import PatternFill
import PIL.Image
from PIL import ImageEnhance, ImageFilter

# --- Config ---
IMAGE_DIR = "beer_images"  # Directory containing beer images
CSV_INPUT_PATH = "beers.csv"  # CSV with scraped beer data and URLs
CSV_OUTPUT_PATH = "beer_label_analysis.csv"  # Output CSV for analysis results
EXCEL_OUTPUT_PATH = "beer_label_analysis_colors.xlsx"  # Output Excel with color cells
MAX_IMAGES = 10  # Limit number of images processed per run (for testing)

# --- Load beer URLs from beers.csv ---
beer_df = pd.read_csv(CSV_INPUT_PATH)

def extract_url(cell):
    """
    Extracts the actual URL from a HYPERLINK formula cell (Excel format) or returns the cell as-is.
    """
    if isinstance(cell, str) and cell.startswith('=HYPERLINK('):
        parts = cell.split('"')
        if len(parts) > 1:
            return parts[1]
    return cell

# Clean up the URL column and create a mapping from image file to URL
beer_df['URL'] = beer_df['URL'].apply(extract_url)
image_to_url = dict(zip(beer_df['Image_File'], beer_df['URL']))

# --- Helper Functions ---
def preprocess_image(image_path):
    """
    Loads and preprocesses an image: resize and denoise for consistency.
    """
    img = cv2.imread(image_path)
    if img is None:
        return None
    img = cv2.resize(img, (400, 400))  # Standardize size
    img = cv2.bilateralFilter(img, 11, 17, 17)  # Denoise
    return img

def segment_bottle(img):
    """
    Segments the bottle/can from a plain white background using thresholding and contour detection.
    Returns the cropped bottle/can region.
    """
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    _, mask = cv2.threshold(gray, 240, 255, cv2.THRESH_BINARY_INV)
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        return img  # fallback: return whole image
    cnt = max(contours, key=cv2.contourArea)
    x, y, w, h = cv2.boundingRect(cnt)
    bottle = img[y:y+h, x:x+w]
    return bottle

def segment_label(bottle_img):
    """
    Attempts to segment the label from the bottle/can using edge detection and contour heuristics.
    Returns the largest plausible label region, or the whole bottle if not found.
    """
    gray = cv2.cvtColor(bottle_img, cv2.COLOR_BGR2GRAY)
    edges = cv2.Canny(gray, 50, 150)
    contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    label_candidates = []
    for cnt in contours:
        x, y, w, h = cv2.boundingRect(cnt)
        aspect_ratio = w / float(h)
        if 1.2 < aspect_ratio < 6 and w > 40 and h > 20:
            label_candidates.append(bottle_img[y:y+h, x:x+w])
    if label_candidates:
        # Return the largest candidate region
        return max(label_candidates, key=lambda im: im.shape[0]*im.shape[1])
    else:
        return bottle_img  # fallback: use whole bottle

def ocr_label(label_img):
    """
    Runs OCR on the label image using EasyOCR and returns the detected text.
    """
    reader = easyocr.Reader(['en'], gpu=False)
    result = reader.readtext(label_img)
    return ' '.join([d[1] for d in result]) if result else ""

def dominant_color(img, k=3):
    """
    Finds the k dominant colors in the image using KMeans clustering.
    Returns a list of RGB tuples.
    """
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    img = img.reshape((-1, 3))
    kmeans = KMeans(n_clusters=k, n_init=10).fit(img)
    colors = kmeans.cluster_centers_.astype(int)
    return [tuple(map(int, c)) for c in colors]

def rgb_to_hex(rgb):
    """
    Converts an RGB tuple to a hex color string for Excel cell coloring.
    """
    return '{:02X}{:02X}{:02X}'.format(*rgb)

def save_colors_to_excel(results, excel_path):
    """
    Saves the analysis results to an Excel file, filling color cells with their actual color and making the beer URL clickable.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Image_File", "OCR_Text", "Label_Color_1", "Label_Color_2", "Label_Color_3", "Bottle_Color_1", "Bottle_Color_2", "Bottle_Color_3", "Beer_URL"])
    for row in results:
        image_file, ocr_text, label_colors, bottle_colors, beer_url = row
        label_colors = eval(label_colors) if isinstance(label_colors, str) else label_colors
        bottle_colors = eval(bottle_colors) if isinstance(bottle_colors, str) else bottle_colors
        excel_row = [image_file, ocr_text]
        for color in label_colors:
            excel_row.append(str(color))
        for color in bottle_colors:
            excel_row.append(str(color))
        excel_row.append(beer_url)
        ws.append(excel_row)
        row_idx = ws.max_row
        # Set cell background colors for label and bottle colors
        for i, color in enumerate(label_colors):
            hex_color = rgb_to_hex(color)
            ws.cell(row=row_idx, column=3+i).fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        for i, color in enumerate(bottle_colors):
            hex_color = rgb_to_hex(color)
            ws.cell(row=row_idx, column=6+i).fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        # Make Beer_URL cell a clickable hyperlink if valid
        if beer_url and isinstance(beer_url, str) and beer_url.startswith('http'):
            ws.cell(row=row_idx, column=9).hyperlink = beer_url
            ws.cell(row=row_idx, column=9).style = 'Hyperlink'
    wb.save(excel_path)

def preprocess_for_ocr(label_img):
    pil_img = PIL.Image.fromarray(cv2.cvtColor(label_img, cv2.COLOR_BGR2RGB))
    pil_img = pil_img.convert('L').filter(ImageFilter.SHARPEN)
    pil_img = ImageEnhance.Contrast(pil_img).enhance(2)
    return cv2.cvtColor(np.array(pil_img), cv2.COLOR_GRAY2BGR)

# --- Main Pipeline ---
results = []
image_count = 0
for image_file in os.listdir(IMAGE_DIR):
    # Only process image files
    if not (image_file.lower().endswith('.jpg') or image_file.lower().endswith('.png') or image_file.lower().endswith('.jpeg')):
        continue
    image_count += 1
    if image_count > MAX_IMAGES:
        break
    image_path = os.path.join(IMAGE_DIR, image_file)
    img = preprocess_image(image_path)
    if img is None:
        # If image can't be loaded, record error and skip
        results.append([image_file, "Error: cannot load image", "", "", ""])
        continue
    # Segment bottle/can from background
    bottle = segment_bottle(img)
    # Segment label from bottle/can
    label = segment_label(bottle)
    # Preprocess label for OCR
    label = preprocess_for_ocr(label)
    # Run OCR on label
    ocr_text = ocr_label(label)
    # Get dominant colors for label and bottle
    label_colors = dominant_color(label)
    bottle_colors = dominant_color(bottle)
    # Get beer URL from beers.csv
    beer_url = image_to_url.get(image_file, "")
    # Append all results for this image
    results.append([
        image_file,
        ocr_text,
        str(label_colors),
        str(bottle_colors),
        beer_url
    ])
    cv2.imwrite(f"debug_label_{image_file}", label)

# --- Save Results ---
# Save to CSV
columns = ["Image_File", "OCR_Text", "Label_Colors", "Bottle_Colors", "Beer_URL"]
df = pd.DataFrame(results, columns=columns)
df.to_csv(CSV_OUTPUT_PATH, index=False)
print(f"✅ Done. Saved analysis to: {CSV_OUTPUT_PATH}")
# Save to Excel with color backgrounds and hyperlinks
save_colors_to_excel(results, EXCEL_OUTPUT_PATH)
print(f"✅ Excel file with color backgrounds saved as {EXCEL_OUTPUT_PATH}")
